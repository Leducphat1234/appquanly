import tkinter as tk
from tkinter import ttk, messagebox, filedialog
from datetime import datetime
try:
    import openpyxl
    from openpyxl.utils import get_column_letter
except ImportError:
    openpyxl = None

class NhanVienTab(ttk.Frame):
    def __init__(self, parent, db):
        super().__init__(parent)
        self.db = db

        notebook = ttk.Notebook(self)
        self.teacher_tab = ttk.Frame(notebook)
        self.class_tab = ttk.Frame(notebook)
        self.session_tab = ttk.Frame(notebook)

        notebook.add(self.teacher_tab, text="Giáo viên")
        notebook.add(self.class_tab, text="Lớp học")
        notebook.add(self.session_tab, text="Buổi dạy")
        notebook.pack(fill='both', expand=True)

        self.build_teacher_tab()
        self.build_class_tab()
        self.build_session_tab()

    def build_teacher_tab(self):
        toolbar = ttk.Frame(self.teacher_tab)
        toolbar.pack(fill='x', padx=10, pady=8)
        ttk.Button(toolbar, text="Thêm", command=lambda: self.open_teacher_form()).pack(side='left', padx=3)
        ttk.Button(toolbar, text="Sửa", command=self.edit_teacher).pack(side='left', padx=3)
        ttk.Button(toolbar, text="Xóa", command=self.delete_teacher).pack(side='left', padx=3)
        ttk.Button(toolbar, text="Xuất Excel", command=self.export_teacher_sessions_report).pack(side='right', padx=3)
        ttk.Button(toolbar, text="Làm mới", command=self.load_teacher_data).pack(side='right', padx=3)

        self.teacher_tree = ttk.Treeview(
            self.teacher_tab,
            columns=("stt", "ten", "luong", "so_lop", "so_tiet", "thu_nhap", "thuong", "id"),
            show="headings"
        )
        headings = [
            ("stt", "STT", 40),
            ("ten", "Họ tên", 220),
            ("luong", "Lương/tiết", 120),
            ("so_lop", "Số lớp", 80),
            ("so_tiet", "Số tiết", 80),
            ("thu_nhap", "Thu nhập", 120),
            ("thuong", "Tiền thưởng", 120),
        ]
        for col, text, width in headings:
            self.teacher_tree.heading(col, text=text)
            self.teacher_tree.column(col, width=width, anchor="center")
        self.teacher_tree.column("id", width=0, stretch=tk.NO)
        self.teacher_tree.pack(expand=True, fill='both', padx=10, pady=5)

        self.load_teacher_data()

    def load_teacher_data(self):
        for item in self.teacher_tree.get_children():
            self.teacher_tree.delete(item)

        teachers = self.db.get_all_giao_vien()
        summary = {row[0]: row[2] for row in self.db.get_giao_vien_summary()}

        for idx, row in enumerate(teachers, start=1):
            tid, name, salary, num_class, bonus = row
            total_sessions = summary.get(tid, 0)
            total_income = total_sessions * salary + bonus
            self.teacher_tree.insert(
                "",
                tk.END,
                values=(
                    idx,
                    name,
                    f"{salary:,.0f}",
                    num_class,
                    total_sessions,
                    f"{total_income:,.0f}",
                    f"{bonus:,.0f}",
                    tid
                )
            )

    def open_teacher_form(self, edit=False):
        selection = None
        if edit:
            selected = self.teacher_tree.selection()
            if selected:
                selection = self.teacher_tree.item(selected[0], "values")

        form = tk.Toplevel(self)
        form.title("Chỉnh sửa giáo viên" if edit else "Thêm giáo viên")
        form.geometry("360x260")
        form.grab_set()

        tk.Label(form, text="Họ tên:").pack(anchor='w', padx=10, pady=(10, 0))
        name_entry = ttk.Entry(form, width=40)
        name_entry.pack(padx=10, pady=5)

        tk.Label(form, text="Lương/tiết:").pack(anchor='w', padx=10, pady=(10, 0))
        salary_entry = ttk.Entry(form, width=40)
        salary_entry.pack(padx=10, pady=5)

        tk.Label(form, text="Tiền thưởng:").pack(anchor='w', padx=10, pady=(10, 0))
        bonus_entry = ttk.Entry(form, width=40)
        bonus_entry.pack(padx=10, pady=5)

        if edit and selection:
            name_entry.insert(0, selection[1])
            salary_entry.insert(0, selection[2].replace(",", ""))
            bonus_entry.insert(0, selection[6].replace(",", ""))

        def save_teacher():
            name = name_entry.get().strip()
            try:
                salary = float(salary_entry.get().strip() or 0)
            except ValueError:
                messagebox.showerror("Lỗi", "Lương/tiết phải là số.")
                return

            try:
                bonus = float(bonus_entry.get().strip() or 0)
            except ValueError:
                messagebox.showerror("Lỗi", "Tiền thưởng phải là số.")
                return

            if not name:
                messagebox.showwarning("Thông báo", "Vui lòng nhập tên giáo viên.")
                return

            if edit and selection:
                real_id = selection[7]
                self.db.cap_nhat_giao_vien(real_id, name, salary, bonus)
            else:
                self.db.them_giao_vien(name, salary, bonus)

            self.load_teacher_data()
            form.destroy()

        ttk.Button(form, text="Lưu", command=save_teacher).pack(pady=15)

    def edit_teacher(self):
        selected = self.teacher_tree.selection()
        if not selected:
            messagebox.showwarning("Thông báo", "Vui lòng chọn giáo viên để sửa")
            return
        self.open_teacher_form(edit=True)

    def delete_teacher(self):
        selected = self.teacher_tree.selection()
        if not selected:
            messagebox.showwarning("Thông báo", "Vui lòng chọn giáo viên để xóa")
            return
        values = self.teacher_tree.item(selected[0], 'values')
        if messagebox.askyesno("Xác nhận", f"Xóa giáo viên {values[1]}?"):
            self.db.xoa_giao_vien(values[7])
            self.load_teacher_data()

    def export_teacher_sessions_report(self):
        if openpyxl is None:
            messagebox.showwarning(
                "Thiếu thư viện",
                "Vui lòng cài đặt openpyxl:\npip install openpyxl"
            )
            return

        filename = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel workbook", "*.xlsx"), ("All files", "*.*")],
            title="Lưu báo cáo buổi dạy"
        )
        if not filename:
            return

        rows = self.db.get_teacher_session_report()
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Báo cáo buổi dạy"
        headers = ["Tên giáo viên", "Số tiết", "Tiền lương", "Tiền thưởng"]
        sheet.append(headers)

        for name, so_tiet, luong, thuong in rows:
            sheet.append([
                name,
                so_tiet,
                float(luong),
                float(thuong)
            ])

        for idx, width in enumerate((30, 12, 15, 15), start=1):
            sheet.column_dimensions[get_column_letter(idx)].width = width

        workbook.save(filename)
        messagebox.showinfo("Thành công", f"Đã xuất báo cáo Excel:\n{filename}")

    def build_class_tab(self):
        toolbar = ttk.Frame(self.class_tab)
        toolbar.pack(fill='x', padx=10, pady=8)
        ttk.Button(toolbar, text="Thêm", command=lambda: self.open_class_form()).pack(side='left', padx=3)
        ttk.Button(toolbar, text="Sửa", command=self.edit_class).pack(side='left', padx=3)
        ttk.Button(toolbar, text="Xóa", command=self.delete_class).pack(side='left', padx=3)
        ttk.Button(toolbar, text="Làm mới", command=self.load_class_data).pack(side='right', padx=3)

        self.class_tree = ttk.Treeview(
            self.class_tab,
            columns=("stt", "ten", "mo_ta", "giao_vien", "id"),
            show="headings"
        )
        headings = [
            ("stt", "STT", 40),
            ("ten", "Tên lớp", 200),
            ("mo_ta", "Mô tả", 250),
            ("giao_vien", "Giáo viên", 200),
        ]
        for col, text, width in headings:
            self.class_tree.heading(col, text=text)
            self.class_tree.column(col, width=width)
        self.class_tree.column("id", width=0, stretch=tk.NO)
        self.class_tree.pack(expand=True, fill='both', padx=10, pady=5)

        self.load_class_data()

    def load_class_data(self):
        for item in self.class_tree.get_children():
            self.class_tree.delete(item)

        classes = self.db.get_all_lop()
        for idx, row in enumerate(classes, start=1):
            cid, ten, mo_ta, gv_id, gv_name = row
            self.class_tree.insert(
                "",
                tk.END,
                values=(idx, ten, mo_ta or "", gv_name or "Chưa có", cid)
            )

    def open_class_form(self, edit=False):
        selection = None
        if edit:
            selected = self.class_tree.selection()
            if not selected:
                messagebox.showwarning("Thông báo", "Vui lòng chọn lớp để sửa")
                return
            selection = self.class_tree.item(selected[0], 'values')

        form = tk.Toplevel(self)
        form.title("Chỉnh sửa lớp" if edit else "Thêm lớp")
        form.geometry("420x300")
        form.grab_set()

        tk.Label(form, text="Tên lớp:").pack(anchor='w', padx=10, pady=(10, 0))
        name_entry = ttk.Entry(form, width=45)
        name_entry.pack(padx=10, pady=5)

        tk.Label(form, text="Mô tả:").pack(anchor='w', padx=10, pady=(10, 0))
        desc_entry = ttk.Entry(form, width=45)
        desc_entry.pack(padx=10, pady=5)

        tk.Label(form, text="Giáo viên:").pack(anchor='w', padx=10, pady=(10, 0))
        teachers = self.db.get_all_giao_vien()
        teacher_options = ["Chưa có"] + [name for (_, name, *_) in teachers]
        teacher_select = ttk.Combobox(form, values=teacher_options, state='readonly', width=42)
        teacher_select.pack(padx=10, pady=5)
        teacher_select.current(0)

        if edit and selection:
            name_entry.insert(0, selection[1])
            desc_entry.insert(0, selection[2])
            if selection[3] != 'Chưa có':
                for idx, (_id, _name, *_) in enumerate(teachers, start=1):
                    if _name == selection[3]:
                        teacher_select.current(idx)
                        break

        def save_class():
            ten_lop = name_entry.get().strip()
            if not ten_lop:
                messagebox.showwarning("Lỗi", "Vui lòng nhập tên lớp.")
                return
            mo_ta = desc_entry.get().strip()
            selected_teacher_id = None
            if teacher_select.current() > 0:
                selected_teacher_id = teachers[teacher_select.current() - 1][0]
            if edit and selection:
                self.db.cap_nhat_lop(selection[4], ten_lop, mo_ta, selected_teacher_id)
            else:
                self.db.them_lop(ten_lop, mo_ta, selected_teacher_id)
            form.destroy()
            self.load_class_data()
            self.load_teacher_data()

        ttk.Button(form, text="Lưu", command=save_class).pack(pady=15)

    def edit_class(self):
        selected = self.class_tree.selection()
        if not selected:
            messagebox.showwarning("Thông báo", "Vui lòng chọn lớp để sửa")
            return
        self.open_class_form(edit=True)

    def delete_class(self):
        selected = self.class_tree.selection()
        if not selected:
            messagebox.showwarning("Thông báo", "Vui lòng chọn lớp để xóa")
            return
        values = self.class_tree.item(selected[0], 'values')
        if messagebox.askyesno("Xác nhận", f"Xóa lớp {values[1]}?"):
            self.db.xoa_lop(values[4])
            self.load_class_data()

    def build_session_tab(self):
        toolbar = ttk.Frame(self.session_tab)
        toolbar.pack(fill='x', padx=10, pady=8)
        ttk.Button(toolbar, text="Thêm", command=self.open_session_form).pack(side='left', padx=3)
        ttk.Button(toolbar, text="Xóa", command=self.delete_session).pack(side='left', padx=3)
        ttk.Button(toolbar, text="Làm mới", command=self.load_session_data).pack(side='right', padx=3)

        self.session_tree = ttk.Treeview(
            self.session_tab,
            columns=("stt", "ngay", "lop", "giao_vien", "id"),
            show="headings"
        )
        headings = [
            ("stt", "STT", 40),
            ("ngay", "Ngày dạy", 150),
            ("lop", "Lớp", 200),
            ("giao_vien", "Giáo viên", 200),
        ]
        for col, text, width in headings:
            self.session_tree.heading(col, text=text)
            self.session_tree.column(col, width=width)
        self.session_tree.column("id", width=0, stretch=tk.NO)
        self.session_tree.pack(expand=True, fill='both', padx=10, pady=5)

        self.load_session_data()

    def load_session_data(self):
        for item in self.session_tree.get_children():
            self.session_tree.delete(item)

        sessions = self.db.get_buoi_day_theo_lop()
        for idx, row in enumerate(sessions, start=1):
            sid, ngay, lop, giao_vien = row
            self.session_tree.insert(
                "",
                tk.END,
                values=(idx, ngay, lop, giao_vien, sid)
            )

    def open_session_form(self):
        form = tk.Toplevel(self)
        form.title("Thêm buổi dạy")
        form.geometry("400x250")
        form.grab_set()

        tk.Label(form, text="Lớp:").pack(anchor='w', padx=10, pady=(10, 0))
        classes = self.db.get_all_lop_names()
        class_options = [name for (_, name) in classes]
        class_select = ttk.Combobox(form, values=class_options, state='readonly', width=45)
        class_select.pack(padx=10, pady=5)

        tk.Label(form, text="Giáo viên:").pack(anchor='w', padx=10, pady=(10, 0))
        teachers = self.db.get_all_teacher_names()
        teacher_options = [name for (_, name) in teachers]
        teacher_select = ttk.Combobox(form, values=teacher_options, state='readonly', width=45)
        teacher_select.pack(padx=10, pady=5)

        tk.Label(form, text="Ngày dạy (YYYY-MM-DD):").pack(anchor='w', padx=10, pady=(10, 0))
        date_entry = ttk.Entry(form, width=45)
        date_entry.pack(padx=10, pady=5)
        date_entry.insert(0, datetime.now().strftime("%Y-%m-%d"))

        def save_session():
            if not class_select.get() or not teacher_select.get() or not date_entry.get():
                messagebox.showwarning("Lỗi", "Vui lòng điền đầy đủ thông tin.")
                return
            
            class_id = [cid for cid, name in classes if name == class_select.get()][0]
            teacher_id = [tid for tid, name in teachers if name == teacher_select.get()][0]
            self.db.them_buoi_day(class_id, teacher_id, date_entry.get())
            self.load_session_data()
            form.destroy()

        ttk.Button(form, text="Lưu", command=save_session).pack(pady=15)

    def delete_session(self):
        selected = self.session_tree.selection()
        if not selected:
            messagebox.showwarning("Thông báo", "Vui lòng chọn buổi dạy để xóa")
            return
        values = self.session_tree.item(selected[0], 'values')
        if messagebox.askyesno("Xác nhận", f"Xóa buổi dạy ngày {values[1]}?"):
            self.cursor.execute("DELETE FROM buoi_day WHERE id=?", (values[4],))
            self.db.conn.commit()
            self.load_session_data()

    def load_data(self):
        """Gọi khi tab được chọn"""
        self.load_teacher_data()
        self.load_class_data()
        self.load_session_data()
