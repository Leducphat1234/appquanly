import tkinter as tk
from tkinter import ttk, messagebox
import os
import calendar
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment

class LichSuTab(ttk.Frame):
    def __init__(self, parent, db):
        super().__init__(parent)
        self.db = db

        # Thanh công cụ
        toolbar = ttk.Frame(self)
        toolbar.pack(fill='x', padx=10, pady=10)
        
        tk.Label(toolbar, text="LỊCH SỬ NGHỈ HỌC", font=("Arial", 12, "bold")).pack(side='left')
        ttk.Button(toolbar, text="Xuất Excel & Xóa dữ liệu cũ", command=self.xuat_excel).pack(side='right')

        # Bảng hiển thị
        self.tree = ttk.Treeview(self, columns=("ngay", "ten"), show="headings")
        self.tree.heading("ngay", text="Ngày vắng")
        self.tree.heading("ten", text="Học sinh")
        self.tree.pack(expand=True, fill='both', padx=10, pady=5)
        
        self.load_data()

    def load_data(self):
        for item in self.tree.get_children(): self.tree.delete(item)
        
        self.db.cursor.execute('''
            SELECT dd.ngay, hv.ho_ten FROM diem_danh dd
            JOIN hoc_vien hv ON dd.id_hoc_vien = hv.id
            WHERE dd.trang_thai = 'Vắng'
            ORDER BY dd.ngay DESC
        ''')
        for row in self.db.cursor.fetchall():
            self.tree.insert("", tk.END, values=(row[0], row[1]))

    def xuat_excel(self):
        # 1. Lấy dữ liệu vắng từ Database
        self.db.cursor.execute('''
            SELECT dd.id, hv.ho_ten, hv.ngay_sinh, dd.ngay, hv.hoc_phi_goc, hv.giam_gia 
            FROM diem_danh dd
            JOIN hoc_vien hv ON dd.id_hoc_vien = hv.id
            WHERE dd.trang_thai = 'Vắng'
            ORDER BY dd.ngay ASC
        ''')
        records = self.db.cursor.fetchall()

        if not records:
            messagebox.showinfo("Thông báo", "Không có dữ liệu vắng mặt để xuất.")
            return

        if not messagebox.askyesno("Xác nhận", "Hệ thống sẽ ghi đè dữ liệu ngày hiện tại vào Excel và dọn dẹp lịch sử cũ. Tiếp tục?"):
            return

        base_dir = "Danh_Sach_Vang"
        
        try:
            # Nhóm dữ liệu theo ngày để xử lý Overwrite một lần cho mỗi ngày
            data_by_date = {}
            for rec in records:
                date_str = rec[3] # Cột dd.ngay
                if date_str not in data_by_date:
                    data_by_date[date_str] = []
                data_by_date[date_str].append(rec)

            for date_str, daily_records in data_by_date.items():
                y, m, d = date_str.split('-')
                folder_path = os.path.join(base_dir, y)
                os.makedirs(folder_path, exist_ok=True)
                file_path = os.path.join(folder_path, f"thang{m}.xlsx")

                # Mở hoặc tạo Workbook
                if os.path.exists(file_path):
                    wb = load_workbook(file_path)
                else:
                    wb = Workbook()
                    wb.remove(wb.active)

                # --- ĐẢM BẢO TẤT CẢ SHEET NGÀY VÀ TỔNG KẾT TỒN TẠI ---
                days_in_month = calendar.monthrange(int(y), int(m))[1]
                for day in range(1, days_in_month + 1):
                    s_name = f"Ngay_{day:02d}"
                    if s_name not in wb.sheetnames:
                        ws_temp = wb.create_sheet(title=s_name)
                        ws_temp.append(["Họ và Tên", "Năm sinh", "Học phí gốc", "Giảm giá (%)"])
                        for cell in ws_temp[1]: cell.font = Font(bold=True)
                
                if "Tong_Ket_Thang" not in wb.sheetnames:
                    summary_ws = wb.create_sheet(title="Tong_Ket_Thang")
                    summary_ws.append(["Họ và Tên", "Tổng số buổi nghỉ"])
                    summary_ws.column_dimensions['A'].width = 25
                    for cell in summary_ws[1]: cell.font = Font(bold=True)

                # --- XỬ LÝ OVERWRITE BẢNG TRONG NGÀY ---
                sheet_name = f"Ngay_{int(d):02d}"
                ws = wb[sheet_name]
                
                # Xóa toàn bộ dữ liệu cũ của ngày đó (giữ lại Header ở dòng 1)
                if ws.max_row > 1:
                    ws.delete_rows(2, ws.max_row)
                
                # Ghi dữ liệu mới vào
                for rec in daily_records:
                    # rec = (id, ho_ten, ngay_sinh, ngay, hp_goc, giam_gia)
                    ws.append([rec[1], rec[2], rec[4], rec[5]])

                # --- CẬP NHẬT SHEET TỔNG KẾT (ĐẾM LẠI TOÀN BỘ) ---
                summary_ws = wb["Tong_Ket_Thang"]
                summary_ws.delete_rows(2, summary_ws.max_row)
                
                counter = {}
                for sn in wb.sheetnames:
                    if sn.startswith("Ngay_"):
                        curr_ws = wb[sn]
                        for row in curr_ws.iter_rows(min_row=2, values_only=True):
                            if row[0]: # Nếu có tên học sinh
                                counter[row[0]] = counter.get(row[0], 0) + 1
                
                for name, total in counter.items():
                    summary_ws.append([name, total])

                wb.save(file_path)

            # 2. Xóa dữ liệu khỏi Database sau khi tất cả file đã lưu xong
            list_ids = [str(r[0]) for r in records]
            placeholders = ",".join("?" * len(list_ids))
            self.db.cursor.execute(f"DELETE FROM diem_danh WHERE id IN ({placeholders})", tuple(list_ids))
            self.db.conn.commit()

            messagebox.showinfo("Thành công", "Đã xuất Excel và cập nhật tổng kết.")
            self.load_data()

        except Exception as e:
            messagebox.showerror("Lỗi", f"Lỗi: {e}\nHãy chắc chắn đã đóng file Excel trước khi bấm nút.")